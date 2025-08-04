import os
import streamlit as st
from openpyxl import load_workbook
from core.utils import get_member_path
from core.search_member import search_member
import logging

logger = logging.getLogger(__name__)

def to_float(value, default=0.0):
    try:
        return float(value)
    except:
        return default

def to_int(value, default=0):
    try:
        return int(value)
    except:
        return default

def list_all_borrowers():
    members_dir = "members"
    try:
        files = [f for f in os.listdir(members_dir) if f.endswith(".xlsx")]
    except OSError as e:
        logger.error(f"Error accessing members directory: {e}")
        st.error(f"❌ Error accessing members directory: {e}")
        return

    if not files:
        st.warning("No borrowers found.")
        return

    st.subheader("📄 All Borrowers Summary")

    for file in files:
        filepath = os.path.join(members_dir, file)
        try:
            wb = load_workbook(filepath)
            ws = wb["Payments"]

            name = ws['A2'].value or "Unknown"
            loan_amount = to_float(ws['B2'].value)
            interest_rate = to_float(ws['C2'].value)
            loan_period = to_int(ws['E2'].value)
            monthly_interest = to_float(ws['F2'].value)

            total_paid = 0
            row = 5
            while ws.cell(row=row, column=7).value:
                try:
                    total_paid += float(ws.cell(row=row, column=8).value)
                except:
                    pass
                row += 1

            total_due = monthly_interest * loan_period
            remaining = total_due - total_paid

            with st.expander(f"👤 {name} — ₹{loan_amount:.2f}"):
                col1, col2 = st.columns([4, 1])
                col1.markdown(f"""
                - 📊 **Loan Period:** {loan_period} months
                - 💸 **Monthly Interest:** ₹{monthly_interest:.2f}
                - ✅ **Interest Paid:** ₹{total_paid:.2f}
                - ⚠️ **Remaining Interest:** ₹{remaining:.2f}
                """)
                if col2.button("🔍 View Profile", key=f"view_{name}_{file}"):
                    st.session_state.current_borrower = name
                    st.session_state.current_view = "🔍 Search Borrower"
                    search_member(name)
        except Exception as e:
            logger.error(f"Error loading data for {file}: {e}")
            st.error(f"❌ Error loading data for {file}: {e}")