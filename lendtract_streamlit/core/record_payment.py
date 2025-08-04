import os
from openpyxl import load_workbook
from core.utils import get_member_path
import logging

logger = logging.getLogger(__name__)

def record_payment(name, amount, date, note=""):
    if amount < 0 or not name.strip():
        logger.error(f"Invalid input: name='{name}', amount={amount}")
        return False

    filename = get_member_path(name)
    if not os.path.exists(filename):
        logger.error(f"File not found: {filename}")
        return False

    try:
        wb = load_workbook(filename)
        ws = wb["Payments"]

        row = 5
        while ws.cell(row=row, column=7).value or ws.cell(row=row, column=8).value:
            row += 1

        ws.cell(row=row, column=7).value = date
        ws.cell(row=row, column=8).value = float(amount)
        ws.cell(row=row, column=9).value = note

        wb.save(filename)
        return True
    except Exception as e:
        logger.error(f"Error recording payment for {name}: {e}")
        return False