import streamlit as st
from datetime import datetime
import os
from openpyxl import load_workbook
from core.utils import get_member_path
from core.record_payment import record_payment
import logging

logger = logging.getLogger(__name__)

def search_member(name):
    filename = get_member_path(name)
    if not os.path.exists(filename):
        st.error("❌ Member not found.")
        return

    try:
        wb = load_workbook(filename)
        ws = wb["Payments"]

        borrower_name = ws['A2'].value
        loan_amount = ws['B2'].value
        interest = ws['C2'].value
        start_date = ws['D2'].value
        loan_period = ws['E2'].value
        monthly_interest = ws['F2'].value

        st.subheader("📄 Borrower Profile")
        st.markdown(f"**👤 Name:** {borrower_name}")
        st.markdown(f"**💰 Loan Amount:** ₹{float(loan_amount):.2f}")
        st.markdown(f"**📈 Interest Rate:** {float(interest):.2f}%")
        st.markdown(f"**📅 Start Date:** {start_date}")
        st.markdown(f"**⏳ Loan Period:** {int(loan_period)} months")
        st.markdown(f"**💸 Monthly Interest:** ₹{float(monthly_interest):.2f}")

        # --- Load Payment History ---
        def load_payment_history():
            row = 5
            history = []
            total_interest_paid = 0

            while row <= ws.max_row:
                payment_date = ws.cell(row=row, column=7).value
                amount_paid = ws.cell(row=row, column=8).value
                note = ws.cell(row=row, column=9).value

                if not payment_date and not amount_paid:
                    break

                try:
                    amount_val = float(amount_paid)
                    total_interest_paid += amount_val
                except:
                    amount_val = 0

                history.append([payment_date, amount_val, note])
                row += 1

            try:
                total_interest_due = float(monthly_interest) * int(loan_period)
                remaining_interest = total_interest_due - total_interest_paid
            except (TypeError, ValueError):
                total_interest_due = 0
                remaining_interest = 0

            return history, total_interest_paid, remaining_interest

        # --- Display Payment History ---
        history, total_interest_paid, remaining_interest = load_payment_history()
        st.markdown("---")
        st.subheader("📜 Payment History")

        if history:
            st.table(history)
            st.success(f"💰 Total Interest Paid: ₹{total_interest_paid:.2f}")
            st.warning(f"📉 Remaining Interest Due: ₹{remaining_interest:.2f}")

            try:
                progress = min(int((total_interest_paid / (float(monthly_interest) * int(loan_period))) * 100), 100)
                st.progress(progress)
            except (ZeroDivisionError, ValueError):
                st.progress(0)
        else:
            st.info("No payments made yet.")

        # --- Record New Payment ---
        st.markdown("---")
        st.subheader("➕ Record New Payment")

        with st.form("add_payment_form"):
            payment_date = st.date_input("Payment Date", value=datetime.today())
            amount_paid = st.number_input("Amount Paid", min_value=0.0)
            note = st.text_input("Note (optional)")
            submitted = st.form_submit_button("Add Payment")

            if submitted:
                success = record_payment(name, amount_paid, str(payment_date), note)
                if success:
                    st.success(f"✅ Payment of ₹{amount_paid:.2f} added on {payment_date}.")
                    st.session_state.current_borrower = name  # Preserve borrower
                    st.session_state.current_view = "🔍 Search Borrower"  # Stay on profile
                    st.rerun()
                else:
                    st.error("❌ Failed to record payment. Check logs for details.")

        # --- Download Borrower File ---
        st.markdown("---")
        with open(filename, "rb") as f:
            st.download_button("📥 Download Borrower File", f, file_name=os.path.basename(filename))
    except Exception as e:
        logger.error(f"Error loading borrower data for {name}: {e}")
        st.error(f"❌ Error loading borrower data: {e}")
