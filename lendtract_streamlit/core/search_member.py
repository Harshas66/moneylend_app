import streamlit as st
from datetime import datetime
import os
from openpyxl import load_workbook
from core.utils import get_member_path
from core.record_payment import record_payment
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import io
import logging

logger = logging.getLogger(__name__)

def search_member(name):
    filename = get_member_path(name)
    if not os.path.exists(filename):
        st.error("❌ Member not found.")
        return

    try:
        wb = load_workbook(filename)
        ws = wb["Payments"]

        borrower_name = ws['A2'].value
        loan_amount = ws['B2'].value
        interest = ws['C2'].value
        start_date = ws['D2'].value
        loan_period = ws['E2'].value
        monthly_interest = ws['F2'].value

        st.subheader("📄 Borrower Profile")
        st.markdown(f"**👤 Name:** {borrower_name}")
        st.markdown(f"**💰 Loan Amount:** ₹{float(loan_amount):.2f}")
        st.markdown(f"**📈 Interest Rate:** {float(interest):.2f}%")
        st.markdown(f"**📅 Start Date:** {start_date}")
        st.markdown(f"**⏳ Loan Period:** {int(loan_period)} months")
        st.markdown(f"**💸 Monthly Interest:** ₹{float(monthly_interest):.2f}")

        def load_payment_history():
            row = 5
            history = []
            total_interest_paid = 0

            while row <= ws.max_row:
                payment_date = ws.cell(row=row, column=7).value
                amount_paid = ws.cell(row=row, column=8).value
                note = ws.cell(row=row, column=9).value

                if not payment_date and not amount_paid:
                    break

                try:
                    amount_val = float(amount_paid)
                    total_interest_paid += amount_val
                except:
                    amount_val = 0

                history.append([payment_date, amount_val, note])
                row += 1

            total_interest_due = float(monthly_interest) * int(loan_period)
            remaining_interest = total_interest_due - total_interest_paid
            return history, total_interest_paid, remaining_interest

        history, total_interest_paid, remaining_interest = load_payment_history()

        st.markdown("---")
        st.subheader("📜 Payment History")

        if history:
            st.table(history)
            st.success(f"💰 Total Interest Paid: ₹{total_interest_paid:.2f}")
            st.warning(f"📉 Remaining Interest Due: ₹{remaining_interest:.2f}")

            progress = min(int((total_interest_paid / (float(monthly_interest) * int(loan_period))) * 100), 100)
            st.progress(progress)
        else:
            st.info("No payments made yet.")

        st.markdown("---")
        st.subheader("➕ Record New Payment")

        form_key = f"add_payment_form_{name}"
        with st.form(form_key):
            payment_date = st.date_input("Payment Date", value=datetime.today())
            amount_paid = st.number_input("Amount Paid", min_value=0.0)
            note = st.text_input("Note (optional)")
            submitted = st.form_submit_button("Add Payment")

            if submitted:
                success = record_payment(name, amount_paid, str(payment_date), note)
                if success:
                    st.success(f"✅ Payment of ₹{amount_paid:.2f} added on {payment_date}.")
                    st.session_state.current_borrower = name
                    st.session_state.current_view = "🔍 Search Borrower"
                    st.rerun()
                else:
                    st.error("❌ Failed to record payment. Check logs for details.")

        st.markdown("---")
        st.subheader("📥 Download Borrower PDF Report")

        def generate_pdf():
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []

            elements.append(Paragraph("Borrower Profile Report", styles['Title']))
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(f"Name: {borrower_name}", styles['Normal']))
            elements.append(Paragraph(f"Loan Amount: ₹{loan_amount}", styles['Normal']))
            elements.append(Paragraph(f"Interest Rate: {interest}%", styles['Normal']))
            elements.append(Paragraph(f"Start Date: {start_date}", styles['Normal']))
            elements.append(Paragraph(f"Loan Period: {loan_period} months", styles['Normal']))
            elements.append(Paragraph(f"Monthly Interest: ₹{monthly_interest}", styles['Normal']))
            elements.append(Spacer(1, 12))
            elements.append(Paragraph("Payment History", styles['Heading2']))

            table_data = [["Date", "Amount Paid", "Note"]] + [
                [str(p[0]), f"₹{p[1]:.2f}", str(p[2] or "")] for p in history
            ]
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)

            elements.append(Spacer(1, 12))
            elements.append(Paragraph(f"Total Interest Paid: ₹{total_interest_paid:.2f}", styles['Normal']))
            elements.append(Paragraph(f"Remaining Interest Due: ₹{remaining_interest:.2f}", styles['Normal']))

            doc.build(elements)
            pdf = buffer.getvalue()
            buffer.close()
            return pdf

        pdf_bytes = generate_pdf()
        st.download_button("📄 Download as PDF", data=pdf_bytes, file_name=f"{borrower_name}_report.pdf", mime="application/pdf")

    except Exception as e:
        logger.error(f"Error loading borrower data for {name}: {e}")
        st.error(f"❌ Error loading borrower data: {e}")