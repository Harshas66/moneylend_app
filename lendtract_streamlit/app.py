import streamlit as st
import os
from openpyxl import load_workbook
from core.addmember import add_member
from core.view_all import list_all_borrowers
from core.search_member import search_member
from core.utils import get_member_path
from core.addmember import add_member
from core.view_all import list_all_borrowers
from core.search_member import search_member


# Set page config with wide layout for better mobile responsiveness
st.set_page_config(page_title="LendTrack", layout="wide")

# Initialize session state
if 'current_view' not in st.session_state:
    st.session_state.current_view = "🏠 Home"
if 'current_borrower' not in st.session_state:
    st.session_state.current_borrower = None

# Create members directory with error handling
try:
    if not os.path.exists("members"):
        os.makedirs("members")
except OSError as e:
    st.error(f"❌ Failed to create members directory: {e}")

# Function to calculate statistics
def get_statistics():
    members_dir = "members"
    total_borrowers = 0
    total_loan_amount = 0.0
    total_interest_paid = 0.0

    try:
        files = [f for f in os.listdir(members_dir) if f.endswith(".xlsx")]
        total_borrowers = len(files)

        for file in files:
            filepath = os.path.join(members_dir, file)
            try:
                wb = load_workbook(filepath)
                ws = wb["Payments"]
                loan_amount = float(ws['B2'].value or 0)
                total_loan_amount += loan_amount

                row = 5
                while ws.cell(row=row, column=8).value:
                    try:
                        total_interest_paid += float(ws.cell(row=row, column=8).value)
                    except:
                        pass
                    row += 1
            except Exception as e:
                st.warning(f"⚠️ Error reading file {file}: {e}")
    except OSError as e:
        st.error(f"❌ Error accessing members directory: {e}")

    return total_borrowers, total_loan_amount, total_interest_paid

# Sidebar menu
menu_options = ["🏠 Home", "➕ Add Member", "📄 View All Borrowers", "🔍 Search Borrower"]
menu = st.sidebar.selectbox("Choose Action", menu_options, 
                           index=menu_options.index(st.session_state.current_view),
                           key="menu_select")

# Update session state with current menu selection
st.session_state.current_view = menu

# Home Page
if menu == "🏠 Home":
    st.title("🏠 LendTrack - Personal Finance Manager")
    st.markdown("""
        **Welcome to LendTrack!**  
        Manage your personal loans with ease. Track borrowers, record payments, and generate reports effortlessly.
    """)

    # Display statistics
    total_borrowers, total_loan_amount, total_interest_paid = get_statistics()
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Borrowers", total_borrowers)
    with col2:
        st.metric("Total Loan Amount", f"₹{total_loan_amount:.2f}")
    with col3:
        st.metric("Total Interest Paid", f"₹{total_interest_paid:.2f}")

    st.markdown("---")
    st.subheader("Quick Actions")
    
    # Action buttons in a responsive grid
    col1, col2, col3 = st.columns([1, 1, 1])
    with col1:
        if st.button("➕ Add New Borrower", use_container_width=True):
            st.session_state.current_view = "➕ Add Member"
            st.rerun()
    with col2:
        if st.button("📄 View All Borrowers", use_container_width=True):
            st.session_state.current_view = "📄 View All Borrowers"
            st.rerun()
    with col3:
        if st.button("🔍 Search Borrower", use_container_width=True):
            st.session_state.current_view = "🔍 Search Borrower"
            st.rerun()

# Add Member Page
elif menu == "➕ Add Member":
    st.title("➕ Add New Borrower")
    st.markdown("Enter details to add a new borrower to LendTrack.")
    name = st.text_input("Borrower Name")
    loan = st.number_input("Loan Amount", min_value=0.0, step=100.0)
    interest = st.number_input("Interest Rate (%)", min_value=0.0, step=0.1)
    period = st.number_input("Loan Period (Months)", min_value=1, step=1)

    if not name.strip():
        st.warning("⚠️ Borrower name cannot be empty.")
    else:
        monthly_interest = (loan * interest) / 100
        total_interest = monthly_interest * period

        if loan and interest and period:
            st.info(f"📌 Monthly Interest: ₹{monthly_interest:.2f}")
            st.info(f"📈 Total Interest Over {int(period)} Months: ₹{total_interest:.2f}")

        if st.button("Add", use_container_width=True):
            if add_member(name, loan, interest, period, monthly_interest):
                st.success(f"✅ Member '{name}' added successfully!")
            else:
                st.error("❗ Member already exists.")

# View All Borrowers Page
elif menu == "📄 View All Borrowers":
    st.title("📄 All Borrowers")
    list_all_borrowers()

# Search Borrower Page
elif menu == "🔍 Search Borrower":
    st.title("🔍 Search Borrower")
    name = st.text_input("Enter Borrower Name")
    if st.button("Search", use_container_width=True):
        if name.strip():
            st.session_state.current_borrower = name
            st.session_state.current_view = "🔍 Search Borrower"
            search_member(name)
        else:
            st.warning("⚠️ Borrower name cannot be empty.")

# Display borrower profile if selected
if st.session_state.current_borrower and menu != "📄 View All Borrowers":
    search_member(st.session_state.current_borrower)